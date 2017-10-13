#!/usr/bin/env python
# coding=utf-8
# encoding=utf8

"""
Creates an excel spreadsheet formatted for Bridge input from a spreadsheet of
lemmatized forms in the format used by autoLemma.py.

Usage:
    python3 convert_lemmata_format.py (greek | latin) import 
        EQUIVALENCETABLE.xlsx
    python3 convert_lemmata_format.py (greek | latin) convert 
        <source-format> <output-format> TARGETFILE.xlsx --col 2 
        --include-ambiguous
"""

import sys, os
from collections import namedtuple
from itertools import combinations
from argparse import ArgumentParser

# networkx.readthedocs.io
from networkx import Graph, NetworkXError, read_gpickle, write_gpickle, \
                     bfs_tree, to_dict_of_dicts

# openpyxl.readthedocs.io
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.reader.excel import SUPPORTED_FORMATS

LEMMATA_COLUMN = 2

Lemma = namedtuple('Lemma', ['display', 'format'])
Lemma.__doc__ = \
    """
    A `Lemma` is a tuple with `display` and `format` coordinates.

    Coordinates:
        display (str): the string which corresponds to the lemma
        format (str): the format of the lemma (e.g. bridge, lalsa)
    """
  
def breadthFirstTraversal(graph, root, center = True):
    """
    Executes a breadth first traversal of the graph `graph` from the central  
    node `root`. 
    
    Parameters:
        graph (networkx.Graph): the graph being traversed
        root (hashable object): the root node of the traversal
        center (bool): if False, do not include `root` in traversal
        
    Yields:
        node (hashable object): a node of `graph`
    """
    # http://stackoverflow.com/questions/35270881/identify-first-node-after-source-node-having-two-neighbors-in-networkx-digraph
    if center:
        yield root
    bfs = bfs_tree(graph, root)
    successors = bfs.successors(root)
    for node in successors:
        yield node
        successors.extend(set(bfs.successors(node)))
    raise StopIteration()
    
def lemmataEquivalenciesInWorkbook(workbook):
    """
    Iterates over Lemmata pairs in the excel file `workbook`.
    
    Parameters:
        workbook (openpyxl.Workbook): the file being read
        
    Yields:
        lemmata_list (List[Lemma]): a list of Lemma display-locations tuples
    """
    formats = []
    column_combinations = []
    for row in workbook.active:
        if not formats:
            formats = [cell.value.lower() for cell in row]  # get format names from first row
            column_combinations = list(combinations(range(len(row)), 2))
            continue
        for a, b in column_combinations:
            if row[a].value is not None and row[b].value is not None:
                yield (Lemma(row[a].value, formats[a]), 
                       Lemma(row[b].value, formats[b]))
    raise StopIteration()
                
def importEquivalenciesFromWorkbookToGraph(workbook, graph):
    """
    Updates the equivalency graph `graph` with the equivalency relations in
    the excel file `workbook`.
    
    Parameters:
        workbook (openpyxl.Workbook): the file being read
        graph (networkx.Graph): a graph of lemmata with edges connecting
            the equivalent ones whose nodes are Lemma tuples
    """
    for pair in lemmataEquivalenciesInWorkbook(workbook):
        graph.add_edge(*pair)

def convertLemma(lemma, output_format, equivalencies, include_ambiguous = False):
    """
    Convert the lemma `lemma` from `lemma.format` to `output_format` using
        equivalency table `equivalencies`.
        
    Parameters:
        lemma (Lemma): the lemma to be converted
        equivalencies (networkx.Graph): a graph of lemmata with edges connecting
            the equivalent ones whose nodes are Lemma tuples
        source_format (str): the format of `lemma`
        output_format (str): the desired format for conversion
        include_ambiguous (bool): whether to return a lemma when there are
            multiple corresponding lemmata in `output_format`
        
    Returns:
        (Lemma): the lemma in `output_format` corresponding to `lemma`
        
    Raises:
        ValueError if no equivalent lemma found, or if multiple equivalent 
            lemmata are found when `include_ambiguous` is False
    """
    equivalent_lemmata = []
    for neighbor in breadthFirstTraversal(equivalencies, lemma, False):
        if neighbor.format == output_format:
            if include_ambiguous:
                return lemma.display
            else: 
                equivalent_lemmata.append(neighbor)
    if len(equivalent_lemmata) == 1:
        return equivalent_lemmata[0]
      
    elif len(equivalent_lemmata) < 1:
        raise ValueError("No {} lemma found for {}"
                         "".format(output_format, lemma.display))
    else:
        raise ValueError("No unique {} lemma found for {}"
                         "".format(output_format, lemma.display))
        
def convertLemmatizedWorkbook(workbook, equivalencies, source_format, 
                              output_format, include_ambiguous = False, 
                              lemmata_column = LEMMATA_COLUMN):
    """
    Create a new lemmatized spreadsheet from `workbook` with the lemmata 
        displayed in `source_format` replaced by those in `output_format`.
    
    Parameters:
        workbook (openpyxl.Workbook): the file being read
        equivalencies (networkx.Graph): a graph of lemmata with edges connecting
            the equivalent ones whose nodes are Lemma tuples
        source_format (str): the format of `lemma`
        output_format (str): the desired format for conversion
        include_ambiguous (bool): whether to return a lemma when there are
            multiple corresponding lemmata in `output_format`
        lemmata_column (int): the column in which lemmata are stored
        
    Returns:
        the mutated openpyxl.Workbook object `workbook`
    """
    ws = workbook.active
    # Write new lemma column label
    ws.cell(row = 1, column = lemmata_column).value = output_format.upper()
    
    lemmata_converted = 0
    for i in range(2, ws.max_row):
        cell = ws.cell(row = i, column = lemmata_column)
        try:
            lemma = Lemma(cell.value, source_format)
            cell.value = convertLemma(lemma, output_format,
                                      equivalencies, include_ambiguous).display
            lemmata_converted += 1
        except (ValueError, NetworkXError):
            # no (unique) equivalent lemma found
            cell.value = "{}: {}".format(lemma.format, lemma.display)
            
    print("{} of {} lemmata converted to {}".format(lemmata_converted, 
                                                    ws.max_row, output_format))
          
    return workbook

if __name__ == '__main__':
    parser = ArgumentParser(description = "Convert the lemma format of a "
                                          "spreadsheet of lemmatized forms.",
                            usage = "\n    %(prog)s (latin | greek) import "
                                    "\n\tEQUIVALENCETABLE.xlsx"
                                    "\n    %(prog)s (latin | greek) convert "
                                    "\n\tTARGETFILE.xlsx <source-format> "
                                    "<output-format> --col 2")
    parser.add_argument('language', choices = ['latin', 'greek'], 
                        metavar = 'language')
    subparsers = parser.add_subparsers(dest = 'command', metavar = 'command')
    subparsers.required = True
    
    import_parser = subparsers.add_parser('import')
    import_parser.add_argument('file_path', metavar = "EQUIVALENCETABLE.xlsx")
    
    convert_parser = subparsers.add_parser('convert')
    convert_parser.add_argument('file_path', metavar = "TARGETFILE.xlsx")
    convert_parser.add_argument('source_format', metavar = "<source-format>")
    convert_parser.add_argument('output_format', metavar = "<output-format>")
    convert_parser.add_argument('-c', '--col', '--lemmata-column', 
                                dest = 'lemmata_column', metavar = "N",
                                type = int, default = LEMMATA_COLUMN)
    convert_parser.add_argument('-a', '--include-ambiguous',
                                dest = 'include_ambiguous', 
                                action = 'store_true')
    
    convert_parser = subparsers.add_parser('export')
    convert_parser.add_argument('--format', default = 'dictionary',
                                choices = ['dictionary'])
    
    args = parser.parse_args()
    
    equivalencies_path = ('{}_lemmata_equivalencies.gpickle'
                          ''.format(args.language))
    equivalencies_len = 0
    print("Loading equivalencies from {}".format(equivalencies_path))
    try:
        equivalencies = read_gpickle(equivalencies_path)
        equivalencies_len = len(equivalencies)
        print("{} lemmata found.".format(equivalencies_len))
    except FileNotFoundError:
        print("File not found. Creating new equivalency graph...")
        equivalencies = Graph()
        
    if args.command == 'export':
        file_path = '{}_lemmata_equivalencies_{}.txt'.format(args.language, 
                                                             args.format)
        print("Printing all lemmata to {}".format(file_path))
        with open(file_path, "wb") as file_opened:
            file_opened.write(str(to_dict_of_dicts(equivalencies)).encode())
        print("Finished.")
        sys.exit()
        
    print("Loading spreadsheet {}".format(args.file_path))
    try:
        read_only = (args.command == 'import')
        workbook = load_workbook(args.file_path, read_only = read_only)
    except FileNotFoundError:
        parser.error("Cannot find file {}".format(args.file_path))
    except InvalidFileException:
        parser.error("Unable to read {}. Supported formats: {}"
                     "".format(args.file_path, ','.join(SUPPORTED_FORMATS)))
    
    if args.command == 'import':
        print("Importing new equivalencies from {}".format(args.file_path))
        importEquivalenciesFromWorkbookToGraph(workbook, equivalencies)
        print("{} lemmata added".format(len(equivalencies) - equivalencies_len))
        print("Saving changes to {}".format(equivalencies_path))
        equivalencies = write_gpickle(equivalencies, equivalencies_path)
        print("Finished.")
        
    elif args.command == 'convert':
        print("Converting lemmata in {}".format(args.file_path))
        convertLemmatizedWorkbook(workbook, equivalencies, args.source_format,
                                  args.output_format, include_ambiguous = False,
                                  lemmata_column = args.lemmata_column)
        
        print("Saving changes to {}".format(args.file_path))
        workbook.save(args.file_path)
        print("Finished.")