#!/usr/bin/env python
# coding=utf-8

"""
Creates an excel spreadsheet formatted for Bridge input from a spreadsheet of
lemmatized forms in the format used by autoLemma.py.

Usage:
    format_lemmatized_text.py [options] [--] <file> ...
    
Arguments:
    <file>                       one or more excel files in autoLemma.py format
    
Options:
    -h --help                    show this help message and exit
    -o <file>, --output <file>   output file (uses input filename by default)
    --append                     append onto an existing output file
    --dir <path>                 prefixed onto file addresses for input/output
    -e, --echo                   print values of output spreadsheets to console
    --remove-duplicates          remove duplicate lemmata in each section
    --skip-formulae              leave cells with formulae empty
    --text-name <title>          text name (for multiple input files)
    --output-sheet <name>        output sheet name [default: Ready to Import]
"""

from sys import exit
from os.path import normpath, splitext, commonprefix, basename
from mimetypes import guess_type
from collections import defaultdict, namedtuple
from itertools import islice

import utils.excel as excel
from autoLemma import OUTPUT_COLUMNS as INPUT_COLUMNS

# https://github.com/docopt/docopt
from docopt import docopt

# openpyxl.readthedocs.io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.reader.excel import SUPPORTED_FORMATS
  
Lemma = namedtuple('Lemma', ['display', 'locations'])
Lemma.__doc__ = \
    """
    A `Lemma` is a tuple with `display` and `locations` coordinates.

    Coordinates:
        display (str): the string which corresponds to the lemma
        locations (List[str]): a list of locations of the word within
            a certain text; of the form 1, 2, 3, ... or 1.1, 1.2, 2.1, ...
    """
    
Column = excel.Column
    # In this script, `find_value` coordinates of Column tuples will be passed
    # Lemma tuples for their first parameter.
    
DISPLAY_FORMULA = "=LOOKUP({},'VOCAB-Text'!A:J,10,FALSE)"
SHORTDEF_FORMULA = "=LOOKUP({},'VOCAB-Text'!A:J,12,FALSE)"
LONGDEF_FORMULA = "=LOOKUP({},'VOCAB-Text'!A:J,13,FALSE)"
LOCALDEF_FORMULA = "=LOOKUP({},'VOCAB-Text'!$B:$K,10,FALSE)"
# NEW_FORMULA = "=IF(COUNTIF('LEMMATA-Vocab'!A:A,{}) = 0, '', {})"

FORMULAE_COLUMNS = [
    "DISPLAY LEMMA", "SHORTDEF", "LONGDEF", "LOCALDEF", "NEW"
]

OUTPUT_COLUMNS = [
    Column("TITLE", 1, 
           lambda lemma, __: lemma.display),
    Column("LOCATION", 2, 
           lambda lemma, __: ", ".join(str(location) for location \
                                                     in lemma.locations)),
    Column("DISPLAY LEMMA", 3,
           lambda __, row: DISPLAY_FORMULA.format(LEMMA_COLUMN_LETTER, row)),
    Column("SHORTDEF", 4,
           lambda __, row: SHORTDEF_FORMULA.format(LEMMA_COLUMN_LETTER, row)),
    Column("LONGDEF", 5,
           lambda __, row: LONGDEF_FORMULA.format(LEMMA_COLUMN_LETTER, row)),
    Column("LOCALDEF", 6,
           lambda __, row: LOCALDEF_FORMULA.format(LEMMA_COLUMN_LETTER, row)),
    Column("NEW", 7,
           lambda __, row: "")
]
  
LEMMA_COLUMN_LETTER = get_column_letter(
    excel.getColumnByName(OUTPUT_COLUMNS, "TITLE").number
)

def lemmataFromLemmatizedWorkbook(workbook, *, include_duplicates=True):
    """
    Extracts a list of Lemma tuples from the excel file `workbook`.
    
    Parameters:
        workbook (openpyxl.Workbook): the file being read
        
    Returns:
        lemmata_list (List[Lemma]): a list of Lemma display-locations tuples
    """
    lemmata_list = []
    final_lemma_dictionary = defaultdict(list)
    sections_dictionary = defaultdict(dict)
    
    lemmata_col = excel.getColumnByName(INPUT_COLUMNS, "TITLE")
    location_col = excel.getColumnByName(INPUT_COLUMNS, "LOCATION")
    section_col = excel.getColumnByName(INPUT_COLUMNS, "SECTION")
    
    def isValidWorksheet(worksheet, verbose=True):
        headers = next(worksheet.rows)
        for column in [lemmata_col, location_col, section_col]:
            if headers[column.number - 1].value != column.name:
                if verbose: 
                    print("Worksheet {worksheet.title} is invalid input; "
                          "expected column {column.number} to be {column.name}"
                          "".format(worksheet=worksheet, column=column))
                return False
        return True
      
    def addLemmataFromWorksheet(worksheet):
        for i, row in enumerate(islice(worksheet, 1, None), start=1):
            if all((cell.value is None for cell in row)):
                print("Stopping at empty row {}".format(i))
                return
            lemma = row[lemmata_col.number - 1].value
            if lemma is None:
                print(row)
                # raise ValueError("Row {} is missing a lemma.".format(i))
                exit("Row {} is missing a lemma.".format(i))
            section = row[section_col.number - 1].value
            if section is None:
                # raise ValueError("Row {} is missing a section.".format(i))
                exit("Row {} is missing a section.".format(i))
            location = row[location_col.number - 1].value
            if include_duplicates:
                final_lemma_dictionary[lemma].append(location)
            elif lemma not in sections_dictionary[section]:
                sections_dictionary[section][lemma] = location
        
    for worksheet in workbook:
        if isValidWorksheet(worksheet): addLemmataFromWorksheet(worksheet)
                
    if not include_duplicates:
        for lemma_dictionary in sections_dictionary.values():
            for display, location in lemma_dictionary.items():
                final_lemma_dictionary[display].append(location)
    for display, locations in final_lemma_dictionary.items():
        lemmata_list.append(Lemma(display, locations))
        
    lemmata_list.sort(key=lambda lemma: lemma.display)
    return lemmata_list
  
def lemmataFromPathList(paths, **kwargs):
    """
    Extracts a list of lemmata as Lemma tuples from the files given by `paths`.
    Note: removes paths from `paths` successively during iteration.
    
    Parameters:
        paths (List[str]): each path must be an address of an excel spreadsheet
        
    Yields:
        lemma (Lemma): the next Lemma tuple to appear in the files in `paths`
    """
    while len(paths) > 0:
        path = paths.pop(0)
        file_type = guess_type(path)[0]
        if file_type != ('application/vnd.openxmlformats-officedocument.'
                         'spreadsheetml.sheet'):
            print("Only excel spreadsheet input is supported: "
                 "{} appears to be {}.".format(path, file_type))
        print("Loading {}".format(path))
        try:
            workbook = load_workbook(path, read_only=True)
            yield from lemmataFromLemmatizedWorkbook(workbook, **kwargs)
        except FileNotFoundError:
            exit("Cannot not find file in path {}".format(path))
        except InvalidFileException:
            exit("Unable to read {}. Supported formats: {}"
                 "".format(path, ','.join(SUPPORTED_FORMATS)))

if __name__ == '__main__':
    args = docopt(__doc__)
    # Find name of text
    if args['--output']:
        output_name, output_extension = splitext(args['--output'])
        if output_extension != 'xlsx':
            print("Note: autoLemma.py only outputs .xlsx files")
    elif args['--text-name']:
        output_name = args['--text-name'] + '_Input'
    else:
        output_name = splitext(basename(commonprefix(args['<file>'])))[0]
        output_name += '_Input'
        if not output_name: output_name = 'Ready to Import'
        
    # Create input/output path
    path_prefix = (normpath(args['--dir']) + '/') if args['--dir'] else ''
    
    # Extract Word tuple iterators from input
    paths = [path_prefix+path for path in args['<file>']]
    try:
        lemmata = lemmataFromPathList(paths, include_duplicates=\
                                      not args['--remove-duplicates'])
    except IOError:
        exit("Error opening one or more files! {!s}".format(args['file']))
     
    # Write grouped lemmata to new spreadsheets
    if args['--skip-formulae']:
        for name in FORMULAE_COLUMNS:
            excel.replaceColumnFunction(OUTPUT_COLUMNS, name, 
                                        lambda __, row: '')
        
    excel.saveDataToSpreadsheet(
        lemmata, OUTPUT_COLUMNS, output_name, path=path_prefix, 
        append=args['--append'], sheet_title=args['--output-sheet'], 
        echo=args['--echo']
    )
    exit()